import numpy as np
import matplotlib.pyplot as plt


def plot_curve(qhist, target, dqlist, ctrlist, delta_ctrlist, diagidx):
    qdatas = np.array(qhist)
    # draw q in a plot
    plt.plot(qdatas[:, 0], marker="+", linestyle="-", label="qpos 0")
    diagidx_ = [idx for idx in diagidx if idx < len(dqlist[0])]
    dqdatas = np.array(dqlist)[:, diagidx_]
    dq_norm = np.linalg.norm(dqdatas, axis=1)
    dq_absavg = np.mean(np.abs(dqdatas), axis=1)
    plt.plot(dq_absavg, label="dq_absavg")
    ctrdatas = np.array(ctrlist)
    plt.plot(ctrdatas[:, 2], linestyle=":", label="ctrl 2")
    plt.plot(ctrdatas[:, 3], linestyle=":", label="ctrl 3")
    # draw delta_ctrl in a plot
    delta_ctrdatas = np.array(delta_ctrlist)
    plt.plot(delta_ctrdatas[:, 2], linestyle="--", label="delta_ctrl 2")
    plt.plot(delta_ctrdatas[:, 3], linestyle="--", label="delta_ctrl 3")
    # show x=0 line
    plt.axhline(y=0, color="r", linestyle="--")
    plt.axhline(y=target[0], color="r", linestyle="--")
    plt.legend()
    # set y axis range
    # plt.ylim(-target[0]*1.5, target[0]*1.5)

    plt.draw()
    plot_img = plt.gcf().canvas.buffer_rgba()
    plot_img = np.array(plot_img)[:, :, :3]

    return plot_img


def save_A_excel(A, filename):
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill

    labels = [
        "root_x",
        "root_y",
        "root_z",
        "root_rx",
        "root_ry",
        "root_rz",
        "left_thigh",
        "left_calf",
        "left_rod",
        "left_wheel",
        "right_thigh",
        "right_calf",
        "right_rod",
        "right_wheel",
        "root_x_v",
        "root_y_v",
        "root_z_v",
        "root_rx_v",
        "root_ry_v",
        "root_rz_v",
        "left_thigh_v",
        "left_calf_v",
        "left_rod_v",
        "left_wheel_v",
        "right_thigh_v",
        "right_calf_v",
        "right_rod_v",
        "right_wheel_v",
    ]
    df = pd.DataFrame(A)
    # insert a column before the first column
    df.insert(0, "", labels)
    # insert a row before the first row
    new_row = pd.DataFrame([[""] + labels], columns=df.columns)
    df = pd.concat([new_row, df], ignore_index=True)
    df.to_excel(filename, index=False, header=False)
    # 打开Excel文件并获取工作表
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # 定义红色加粗字体样式
    red_bold_font = Font(color="FF0000", bold=True)

    # 定义红色边框样式
    red_border = Border(
        left=Side(style="thin", color="FF0000"),
        right=Side(style="thin", color="FF0000"),
        top=Side(style="thin", color="FF0000"),
        bottom=Side(style="thin", color="FF0000"),
    )

    # 定义灰色背景填充样式
    gray_fill = PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid")
    # 定义浅灰色背景填充样式
    light_gray_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )

    # 遍历所有单元格并应用条件格式
    for row in ws.iter_rows(
        min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)) and np.abs(cell.value) > 1e-3:
                cell.font = red_bold_font
                cell.border = red_border

    # 将第一行和第一列的背景设置为灰色
    for cell in ws[1]:
        cell.fill = gray_fill
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row[0].fill = gray_fill

    # 将B2到O15的单元格设置为浅灰色背景
    for row in ws.iter_rows(min_row=2, max_row=15, min_col=2, max_col=15):
        for cell in row:
            cell.fill = light_gray_fill
    # 将P16到AC29的单元格设置为浅灰色背景
    for row in ws.iter_rows(min_row=16, max_row=29, min_col=16, max_col=29):
        for cell in row:
            cell.fill = light_gray_fill

    # 保存修改后的Excel文件
    wb.save(filename)


def save_B_excel(B, filename):
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill

    labels_col = [
        "root_x",
        "root_y",
        "root_z",
        "root_rx",
        "root_ry",
        "root_rz",
        "left_thigh",
        "left_calf",
        "left_rod",
        "left_wheel",
        "right_thigh",
        "right_calf",
        "right_rod",
        "right_wheel",
        "root_x_v",
        "root_y_v",
        "root_z_v",
        "root_rx_v",
        "root_ry_v",
        "root_rz_v",
        "left_thigh_v",
        "left_calf_v",
        "left_rod_v",
        "left_wheel_v",
        "right_thigh_v",
        "right_calf_v",
        "right_rod_v",
        "right_wheel_v",
    ]
    labels_row = [
        "left_actuator_thigh",
        "right_actuator_thigh",
        "left_actuator_wheel",
        "right_actuator_wheel",
    ]
    df = pd.DataFrame(B)
    # insert a column before the first column
    df.insert(0, "", labels_col)
    # insert a row before the first row
    new_row = pd.DataFrame([[""] + labels_row], columns=df.columns)
    df = pd.concat([new_row, df], ignore_index=True)
    df.to_excel(filename, index=False, header=False)
    # 打开Excel文件并获取工作表
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # 定义红色加粗字体样式
    red_bold_font = Font(color="FF0000", bold=True)

    # 定义红色边框样式
    red_border = Border(
        left=Side(style="thin", color="FF0000"),
        right=Side(style="thin", color="FF0000"),
        top=Side(style="thin", color="FF0000"),
        bottom=Side(style="thin", color="FF0000"),
    )

    # 定义灰色背景填充样式
    gray_fill = PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid")
    # 定义浅灰色背景填充样式
    light_gray_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )

    # 遍历所有单元格并应用条件格式
    for row in ws.iter_rows(
        min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)) and np.abs(cell.value) > 1e-3:
                cell.font = red_bold_font
                cell.border = red_border

    # 将第一行和第一列的背景设置为灰色
    for cell in ws[1]:
        cell.fill = gray_fill
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row[0].fill = gray_fill

    # 将B2到E15的单元格设置为浅灰色背景
    for row in ws.iter_rows(min_row=2, max_row=15, min_col=2, max_col=5):
        for cell in row:
            cell.fill = light_gray_fill

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # 保存修改后的Excel文件
    wb.save(filename)


def plot_smallest_vertical_force(height_offsets, vertical_forces, best_offset, model):
    # Plot the relationship.
    plt.figure(figsize=(10, 6))
    plt.plot(height_offsets * 1000, vertical_forces, linewidth=3)
    # Red vertical line at offset corresponding to smallest vertical force.
    plt.axvline(x=best_offset * 1000, color="red", linestyle="--")
    # Green horizontal line at the humanoid's weight.
    weight = model.body_subtreemass[1] * np.linalg.norm(model.opt.gravity)
    plt.axhline(y=weight, color="green", linestyle="--")
    plt.xlabel("Height offset (mm)")
    plt.ylabel("Vertical force (N)")
    plt.grid(which="major", color="#DDDDDD", linewidth=0.8)
    plt.grid(which="minor", color="#EEEEEE", linestyle=":", linewidth=0.5)
    plt.minorticks_on()
    plt.title(f"Smallest vertical force " f"found at offset {best_offset*1000:.4f}mm.")
    plt.show()
